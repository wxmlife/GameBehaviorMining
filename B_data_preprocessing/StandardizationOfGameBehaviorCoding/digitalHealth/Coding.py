import pandas as pd
import os
import urllib.parse
from openpyxl import load_workbook

def decode_url_encoded_text(text):
    """将URL编码的文本转换为正常文字"""
    try:
        # 将%XX格式的编码转换为正常字符
        decoded = urllib.parse.unquote(text)
        return decoded
    except Exception as e:
        print(f"解码失败: {text}, 错误: {e}")
        return text  # 返回原始文本以防出错

def process_excel_files(input_folder, output_folder):
    """
    批量处理Excel文件中的URL编码文本
    :param input_folder: 输入Excel文件所在的文件夹路径
    :param output_folder: 输出文件夹路径
    """
    # 确保输出文件夹存在
    os.makedirs(output_folder, exist_ok=True)
    
    # 遍历输入文件夹中的所有Excel文件
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)
            
            # 使用openpyxl读取Excel文件
            wb = load_workbook(input_path)
            
            # 遍历所有工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 遍历工作表中的所有单元格
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and '%' in cell.value:
                            # 解码URL编码的文本
                            cell.value = decode_url_encoded_text(cell.value)
            
            # 保存处理后的文件
            wb.save(output_path)
            print(f"已处理文件: {filename}")

# 使用示例
if __name__ == "__main__":
    # 配置参数
    INPUT_FOLDER = './../../../A_data_input/GameBehavior/数字健康_解救计划'
    OUTPUT_FOLDER = "result"    # 替换为输出文件夹路径
    
    # 执行处理
    process_excel_files(INPUT_FOLDER, OUTPUT_FOLDER)
    print("所有文件处理完成！")